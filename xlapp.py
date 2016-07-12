from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color
import os


class Window(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master = master
        self.init_window()
        self.count  = False

    def init_window(self):
        self.master.title("")
        self.pack(fill=BOTH, expand=1)

        L1 = Label(self, text="Wybierz plik:")
        B1 = Button(self, text="Wybierz", command=self.check_for_file)
        self.E1 = Entry(self, width=50)

        L2 = Label(self, text="Odznacz indeks:")
        self.E2 = Entry(self, width=20)
        B2 = Button(self, text="Odznacz", command=self.update_excel)

        L1.place(x=20, y=20)
        B1.place(x=450, y=45)
        self.E1.place(x=20, y=50)
        L2.place(x=20, y=130)
        self.E2.place(x=20, y=165)
        B2.place(x=200, y=160)

    def check_for_file(self):
        try:
            self.file_path = filedialog.askopenfilename()
            self.E1.insert(0, self.file_path)
        except Exception as e:
            messagebox.showerror(title="Uwaga", message="Bledna sciezka dostepu do pliku!")
            messagebox.showerror(title=None, message=e)

    def update_excel(self):
        if '__kopia__pliku__.xlsx' not in self.file_path and self.count == True:
            messagebox.showwarning(title=None,
                                   message="Dla kolejnych odznaczen ustaw sciezke dla nowo stworzonego pliku kopii!")
        else:
            try:
                if self.E2.get().isdigit():
                    find_value = int(self.E2.get())
                else:
                    find_value = self.E2.get()
                book = load_workbook(self.file_path)
                # color_fill = PatternFill(patternType='solid', fgColor=Color('56FFE9'))
                orange = PatternFill(patternType='solid', fgColor=Color('FFFFC000'))
                sheets = book.get_sheet_names()
                check_amount = 0
                for count, sheet in enumerate(sheets):
                    sheet = book.get_sheet_by_name(sheets[count])
                    for row in sheet.rows:
                        for cell in row:
                            if cell.value == find_value:
                                check_amount += 1
                                cell.fill = orange
                if check_amount == 0:
                    messagebox.showinfo(title=None, message="Odznaczenie nie powiodło sie")
                else:
                    book.save('__kopia__pliku__.xlsx')
                    self.count = True
                    messagebox.showinfo(title=None, message="Pomyslnie zaktualizowano dokument")
            except:
                messagebox.showwarning(title="Uwaga", message="Sprobuj ponownie wybrac plik!")


def Main():
    root = Tk()
    e = Entry(root)
    root.geometry("600x300")
    root.resizable(width=False, height=False)
    app = Window(root)
    root.mainloop()


if __name__ == '__main__':
    Main()
